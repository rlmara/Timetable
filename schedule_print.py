from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment

class TT:
	def __init__(self, ppd, dpw, cl, ra, relevance):
		self.ppd = ppd
		self.dpw = dpw
		self.cl = cl
		self.ra = ra
		self.relevance = relevance
		self.wb = Workbook()
		self.tmp = self.wb.active
		self.ws = None

	def fill(self, name, values):
		self.ws = self.wb.create_sheet(str(name))
		r = self.ppd
		c = 1
#		for i in range(len(values)):
		for i in range(self.ppd * self.dpw * self.cl):
			d = self.ws.cell(row=r, column=c, value=values[i])
			# if (c == self.cl * self.dpw):
			# 	c = 1
			# 	r = r - 1
			# else:
			# 	c = c + 1
			if (r == 1):
				c = c + 1
				r = self.ppd
			else:
				r = r - 1

			d.font = Font(bold=True, size=16, name='Calibri', italic=False, vertAlign=None, underline='none', strike=False, color='FFFFFF')
			d.fill = PatternFill("solid", fgColor="000000")
			#d.bd = Side(style='thick', color="000000")
			#d.border = Border(left=bd, top=bd, right=bd, bottom=bd)

		# Make room for headers
		self.ws.move_range("A1:Z20", rows=1, cols=1)
		# Add headers

		for i in range(self.cl * self.dpw):
			c = self.ws.cell(row= 1, column= i+2, value= "Day " + str(i))
			c.font = Font(bold=True, size=12, color='0000FF')
		for j in range(self.ppd, 0, -1):
			c = self.ws.cell(row= j+1, column= 1, value= "Per " + str(self.ppd-j))
			c.font = Font(bold=True, size=12, color='0000FF')

		# Add recess
		self.recess = self.ppd - self.ra +1
		self.ws.insert_rows(self.recess)
		r = self.ws.cell(row=(self.recess), column= 2, value="RECESS")
		r.alignment = Alignment(horizontal='center')
		self.ws.merge_cells(start_row=(self.recess), start_column=2, end_row=(self.recess), end_column=(self.cl * self.dpw)+1)

	def saveNClose(self):
		self.wb.remove(self.tmp)

		for sheet in self.wb.worksheets:
			for row in range(self.ppd + 1):
				if row not in [1, self.recess]:
					sheet.row_dimensions[row].width = 90

			for col in range(self.cl * self.dpw):
				sheet.column_dimensions[get_column_letter(col+2)].width = 20

		if self.relevance == "G":
	   		self.wb.save(filename = 'Output\\group.xlsx')
		if self.relevance == "T":
	   		self.wb.save(filename = 'Output\\teacher.xlsx')
		if self.relevance == "R":
	   		self.wb.save(filename = 'Output\\room.xlsx')
